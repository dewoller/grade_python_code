"""
Excel mock output generator for testing output formatting.

This module creates expected Excel outputs for:
- Perfect student results
- Students with missing tasks
- Students with errors
- Edge cases and formatting tests

All outputs match the expected Excel format from the project brief.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class ExcelMockGenerator:
    """Generate expected Excel output fixtures for testing."""
    
    def __init__(self, fixtures_dir: str = "tests/fixtures"):
        self.fixtures_dir = Path(fixtures_dir)
        self.excel_mocks_dir = self.fixtures_dir / "excel_outputs"
        self.excel_mocks_dir.mkdir(parents=True, exist_ok=True)
    
    def create_perfect_student_excel_data(self) -> List[Dict[str, Any]]:
        """Create Excel data for perfect student."""
        return [
            # Task 2 - 8 points
            {"Task Name": "Task 2", "Criterion": "Data loading implementation", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 2", "Criterion": "Data cleaning functionality", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 2", "Criterion": "Statistical analysis correctness", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 2", "Criterion": "Code organization and structure", "Score": 1, "Max Points": 1},
            {"Task Name": "Task 2", "Criterion": "Documentation and comments", "Score": 1, "Max Points": 1},
            {"Task Name": "Task 2 Subtotal", "Criterion": "", "Score": 8, "Max Points": 8},
            
            # Task 3 - 10 points
            {"Task Name": "Task 3", "Criterion": "Histogram creation and formatting", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 3", "Criterion": "Scatter plot implementation", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 3", "Criterion": "Plot customization and labels", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 3", "Criterion": "Data visualization best practices", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 3", "Criterion": "Code quality and efficiency", "Score": 1, "Max Points": 1},
            {"Task Name": "Task 3", "Criterion": "Error handling for plots", "Score": 1, "Max Points": 1},
            {"Task Name": "Task 3 Subtotal", "Criterion": "", "Score": 10, "Max Points": 10},
            
            # Task 4 - 15 points
            {"Task Name": "Task 4", "Criterion": "Data preprocessing implementation", "Score": 3, "Max Points": 3},
            {"Task Name": "Task 4", "Criterion": "Feature scaling correctness", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 4", "Criterion": "Clustering algorithm application", "Score": 3, "Max Points": 3},
            {"Task Name": "Task 4", "Criterion": "Parameter selection and tuning", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 4", "Criterion": "Results interpretation", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 4", "Criterion": "Code modularity and reusability", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 4", "Criterion": "Documentation and testing", "Score": 1, "Max Points": 1},
            {"Task Name": "Task 4 Subtotal", "Criterion": "", "Score": 15, "Max Points": 15},
            
            # Task 5 - 15 points
            {"Task Name": "Task 5", "Criterion": "Correlation analysis implementation", "Score": 3, "Max Points": 3},
            {"Task Name": "Task 5", "Criterion": "Statistical test selection", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 5", "Criterion": "Hypothesis testing correctness", "Score": 3, "Max Points": 3},
            {"Task Name": "Task 5", "Criterion": "Results interpretation and significance", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 5", "Criterion": "Statistical assumptions validation", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 5", "Criterion": "Code efficiency and clarity", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 5", "Criterion": "Documentation and explanations", "Score": 1, "Max Points": 1},
            {"Task Name": "Task 5 Subtotal", "Criterion": "", "Score": 15, "Max Points": 15},
            
            # Task 6 - 15 points
            {"Task Name": "Task 6", "Criterion": "Model selection and implementation", "Score": 3, "Max Points": 3},
            {"Task Name": "Task 6", "Criterion": "Cross-validation methodology", "Score": 3, "Max Points": 3},
            {"Task Name": "Task 6", "Criterion": "Performance metrics calculation", "Score": 3, "Max Points": 3},
            {"Task Name": "Task 6", "Criterion": "Model evaluation and comparison", "Score": 3, "Max Points": 3},
            {"Task Name": "Task 6", "Criterion": "Code organization and testing", "Score": 2, "Max Points": 2},
            {"Task Name": "Task 6", "Criterion": "Documentation and interpretation", "Score": 1, "Max Points": 1},
            {"Task Name": "Task 6 Subtotal", "Criterion": "", "Score": 15, "Max Points": 15},
            
            # Task 7 - 20 points
            {"Task Name": "Task 7", "Criterion": "Comprehensive report generation", "Score": 5, "Max Points": 5},
            {"Task Name": "Task 7", "Criterion": "Data visualization integration", "Score": 4, "Max Points": 4},
            {"Task Name": "Task 7", "Criterion": "Statistical summary and insights", "Score": 4, "Max Points": 4},
            {"Task Name": "Task 7", "Criterion": "Conclusions and recommendations", "Score": 4, "Max Points": 4},
            {"Task Name": "Task 7", "Criterion": "Professional presentation quality", "Score": 3, "Max Points": 3},
            {"Task Name": "Task 7 Subtotal", "Criterion": "", "Score": 20, "Max Points": 20},
            
            # Grand Total
            {"Task Name": "TOTAL", "Criterion": "", "Score": 83, "Max Points": 83}
        ]
    
    def create_missing_tasks_excel_data(self) -> List[Dict[str, Any]]:
        """Create Excel data for student with missing tasks."""
        data = self.create_perfect_student_excel_data()
        
        # Mark Task 4 as missing (set all scores to 0 with MISSING_TASK flag)
        for row in data:
            if row["Task Name"] == "Task 4" and row["Task Name"] != "Task 4 Subtotal":
                row["Score"] = "0 (MISSING_TASK)"
            elif row["Task Name"] == "Task 4 Subtotal":
                row["Score"] = "0 (MISSING_TASK)"
        
        # Mark Task 6 as missing
        for row in data:
            if row["Task Name"] == "Task 6" and row["Task Name"] != "Task 6 Subtotal":
                row["Score"] = "0 (MISSING_TASK)"
            elif row["Task Name"] == "Task 6 Subtotal":
                row["Score"] = "0 (MISSING_TASK)"
        
        # Update total (83 - 15 - 15 = 53)
        for row in data:
            if row["Task Name"] == "TOTAL":
                row["Score"] = 53
        
        return data
    
    def create_syntax_errors_excel_data(self) -> List[Dict[str, Any]]:
        """Create Excel data for student with syntax errors."""
        data = self.create_perfect_student_excel_data()
        
        # Mark some criteria with parsing errors
        error_criteria = [
            ("Task 2", "Data loading implementation"),
            ("Task 3", "Histogram creation and formatting"),
            ("Task 5", "Correlation analysis implementation")
        ]
        
        for row in data:
            if (row["Task Name"], row["Criterion"]) in error_criteria:
                row["Score"] = "0 (PARSING_ERROR)"
        
        # Update subtotals and total accordingly
        # Task 2: 8 - 2 = 6
        # Task 3: 10 - 2 = 8  
        # Task 5: 15 - 3 = 12
        # Total: 83 - 7 = 76
        
        for row in data:
            if row["Task Name"] == "Task 2 Subtotal":
                row["Score"] = 6
            elif row["Task Name"] == "Task 3 Subtotal":
                row["Score"] = 8
            elif row["Task Name"] == "Task 5 Subtotal":
                row["Score"] = 12
            elif row["Task Name"] == "TOTAL":
                row["Score"] = 76
        
        return data
    
    def create_excel_with_issues_summary(self, data: List[Dict], issues: List[str]) -> pd.DataFrame:
        """Create Excel with issues summary section."""
        df = pd.DataFrame(data)
        
        # Add blank rows and issues summary
        blank_row = {"Task Name": "", "Criterion": "", "Score": "", "Max Points": ""}
        issues_header = {"Task Name": "ISSUES FOUND:", "Criterion": "", "Score": "", "Max Points": ""}
        
        df = pd.concat([df, pd.DataFrame([blank_row, issues_header])], ignore_index=True)
        
        for issue in issues:
            issue_row = {"Task Name": f"- {issue}", "Criterion": "", "Score": "", "Max Points": ""}
            df = pd.concat([df, pd.DataFrame([issue_row])], ignore_index=True)
        
        return df
    
    def save_excel_fixture(self, data: List[Dict], filename: str, issues: List[str] = None):
        """Save Excel fixture with proper formatting."""
        if issues:
            df = self.create_excel_with_issues_summary(data, issues)
        else:
            df = pd.DataFrame(data)
        
        filepath = self.excel_mocks_dir / f"{filename}.xlsx"
        
        # Create Excel with formatting
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Marking Sheet', index=False)
            
            # Get workbook and worksheet for formatting
            workbook = writer.book
            worksheet = writer.sheets['Marking Sheet']
            
            # Apply formatting
            self.apply_excel_formatting(worksheet, len(df))
    
    def apply_excel_formatting(self, worksheet, num_rows: int):
        """Apply formatting to Excel worksheet."""
        # Header formatting
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        
        # Subtotal row formatting
        subtotal_font = Font(bold=True)
        subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        for row_num in range(2, num_rows + 2):
            cell_value = worksheet[f'A{row_num}'].value
            if cell_value and ('Subtotal' in str(cell_value) or cell_value == 'TOTAL'):
                for col in range(1, 5):
                    cell = worksheet.cell(row=row_num, column=col)
                    cell.font = subtotal_font
                    if cell_value == 'TOTAL':
                        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.fill = subtotal_fill
        
        # Issues section formatting
        for row_num in range(2, num_rows + 2):
            cell_value = worksheet[f'A{row_num}'].value
            if cell_value and 'ISSUES FOUND' in str(cell_value):
                worksheet[f'A{row_num}'].font = Font(bold=True, color="C5504B")
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def generate_all_excel_fixtures(self):
        """Generate all Excel mock fixtures."""
        # Perfect student
        perfect_data = self.create_perfect_student_excel_data()
        self.save_excel_fixture(perfect_data, "perfect_student_marks")
        
        # Missing tasks student
        missing_data = self.create_missing_tasks_excel_data()
        missing_issues = [
            "Task 4: MISSING_TASK",
            "Task 6: MISSING_TASK", 
            "Manual review required for 2 items"
        ]
        self.save_excel_fixture(missing_data, "missing_tasks_student_marks", missing_issues)
        
        # Syntax errors student
        syntax_data = self.create_syntax_errors_excel_data()
        syntax_issues = [
            "Task 2 Criterion 1: PARSING_ERROR",
            "Task 3 Criterion 1: PARSING_ERROR",
            "Task 5 Criterion 1: PARSING_ERROR",
            "Manual review required for 3 items"
        ]
        self.save_excel_fixture(syntax_data, "syntax_errors_student_marks", syntax_issues)
        
        # Empty notebook (all zeros)
        empty_data = []
        for row in self.create_perfect_student_excel_data():
            new_row = row.copy()
            if row["Task Name"] not in ["TOTAL"] and "Subtotal" not in row["Task Name"]:
                new_row["Score"] = "0 (MISSING_TASK)"
            else:
                new_row["Score"] = 0
            empty_data.append(new_row)
        
        empty_issues = [
            "All tasks: MISSING_TASK",
            "Complete notebook review required"
        ]
        self.save_excel_fixture(empty_data, "empty_notebook_marks", empty_issues)
